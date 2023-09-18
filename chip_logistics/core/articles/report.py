"""CSV report generation."""


from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from chip_logistics.models.articles import ArticleItem

# Multiplier of content size to adjust column width
COLUMN_WIDTH_RATIO = 1.3


def generate_report_name() -> str:
    """Generate report file name from current datetime.

    Returns:
        Report file name with .xlsx extension.
    """
    return 'Расчет-{date}.xlsx'.format(
        date=datetime.now().strftime('%H:%M-%m.%d.%Y'),
    )


def add_header(sheet: Worksheet, columns_names: list[str]) -> None:
    """Add header with bold columns names to sheet.

    Args:
        sheet: Target worksheet.
        columns_names: Names of columns in header.
    """
    header_font = Font(bold=True)
    for column, column_name in enumerate(columns_names):
        cell = sheet.cell(1, column + 1, value=column_name)
        cell.font = header_font


def adjust_columns_width(sheet: Worksheet) -> None:
    """Adjust columns width to content.

    Args:
        sheet: Worksheet to adjust.
    """
    for column_num, cells in enumerate(sheet.columns, start=1):
        contents_width = (len(str(cell.value)) for cell in cells)
        max_width = max(contents_width, default=0)
        column = sheet.column_dimensions[get_column_letter(column_num)]
        column.width = max_width * COLUMN_WIDTH_RATIO


def create_calculations_report(
    calculations_results: list[tuple[ArticleItem, Decimal]],
    total_price: Decimal,
) -> tuple[bytes, str]:
    """Generate Excel calculations report.

    Args:
        calculations_results: List with item sand their costs.
        total_price: Total items price.

    Returns:
        File data and name.
    """
    workbook = Workbook()
    sheet = workbook.active

    add_header(
        sheet,
        [
            'Наименование',
            'Количество',
            'Общий вес',
            'Таможенная пошлина',
            'Цена',
        ],
    )

    for article_item, price in calculations_results:
        sheet.append([
            article_item.name,
            article_item.count,
            article_item.unit_weight * article_item.count,
            1 - article_item.duty_fee_ratio,
            price,
        ])

    sheet.append([])
    sheet.append(['Общая стоимость', total_price])

    adjust_columns_width(sheet)

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    return file_buffer.getvalue(), generate_report_name()
